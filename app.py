import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import re
from textblob import TextBlob
from spellchecker import SpellChecker

# Configure the page
st.set_page_config(
    page_title="Excel Spell & Grammar Checker",
    page_icon="📝",
    layout="wide"
)

# Initialize spell checker
@st.cache_resource
def load_spell_checker():
    """Load the spell checker"""
    try:
        return SpellChecker()
    except Exception as e:
        st.error(f"Error loading spell checker: {e}")
        return None

def is_text_content(value):
    """Check if a cell contains text content worth checking"""
    if pd.isna(value) or value == "":
        return False
    
    # Convert to string and check if it's meaningful text
    text = str(value).strip()
    
    # Skip if it's just numbers, dates, or very short text
    if len(text) < 2:
        return False
    
    # Skip if it's purely numeric
    if re.match(r'^[\d\.\,\-\+\%\$\€\£\¥]+$', text):
        return False
    
    # Skip if it's a date format
    if re.match(r'^\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}$', text):
        return False
    
    return True

def check_spelling_and_grammar(text, spell_checker):
    """Check text for spelling and grammar issues"""
    issues = []
    
    # Clean the text for checking
    text_clean = re.sub(r'[^\w\s\'\-]', ' ', text)
    words = text_clean.lower().split()
    
    # Spelling check with pyspellchecker
    if spell_checker:
        try:
            misspelled = spell_checker.unknown(words)
            if misspelled:
                issues.append(f"Possible misspelled words: {', '.join(list(misspelled)[:3])}")
        except Exception:
            pass
    
    # Basic grammar check with TextBlob
    try:
        blob = TextBlob(text)
        
        # Check for very basic grammar issues
        # Multiple consecutive spaces
        if re.search(r'\s{2,}', text):
            issues.append("Multiple consecutive spaces found")
        
        # Missing capitalization at start of sentence
        sentences = blob.sentences
        for sentence in sentences:
            sentence_str = str(sentence).strip()
            if sentence_str and sentence_str[0].islower():
                # Check if it's not an intentional lowercase start (like an acronym)
                if len(sentence_str.split()) > 1:
                    issues.append("Sentence may need capitalization")
                break  # Only report once per cell
        
        # Check for repeated words
        words_in_text = text.lower().split()
        for i in range(len(words_in_text) - 1):
            if words_in_text[i] == words_in_text[i + 1] and len(words_in_text[i]) > 2:
                issues.append("Repeated word detected")
                break  # Only report once per cell
                
    except Exception:
        pass
    
    return issues

def process_workbook(uploaded_file, spell_checker):
    """Process the Excel workbook and highlight cells with issues"""
    
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file, data_only=False)
    
    # Yellow fill for highlighting issues
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    issues_found = []
    total_checked = 0
    
    # Process each worksheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        st.write(f"Processing sheet: **{sheet_name}**")
        progress_bar = st.progress(0)
        
        # Get all cells with data
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        cell_count = 0
        total_cells = max_row * max_col
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Update progress
                cell_count += 1
                if cell_count % 100 == 0:  # Update every 100 cells
                    progress_bar.progress(min(cell_count / total_cells, 1.0))
                
                if cell.value and is_text_content(cell.value):
                    total_checked += 1
                    text = str(cell.value).strip()
                    
                    # Check for issues
                    issues = check_spelling_and_grammar(text, spell_checker)
                    
                    if issues:
                        # Highlight the cell
                        cell.fill = yellow_fill
                        
                        # Add comment with issues
                        if cell.comment:
                            existing_comment = cell.comment.text
                            cell.comment.text = f"{existing_comment}\n\nSpell/Grammar Issues:\n" + "\n".join(issues)
                        else:
                            cell.comment = openpyxl.comments.Comment(
                                text=f"Spell/Grammar Issues:\n" + "\n".join(issues),
                                author="Spell Checker"
                            )
                        
                        issues_found.append({
                            'Sheet': sheet_name,
                            'Cell': f"{openpyxl.utils.get_column_letter(col)}{row}",
                            'Text': text,
                            'Issues': "; ".join(issues)
                        })
        
        progress_bar.progress(1.0)
    
    return workbook, issues_found, total_checked

def main():
    st.title("📝 Excel Spell & Grammar Checker")
    st.markdown("Upload an Excel workbook to check for spelling and grammar issues across all sheets.")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls'],
        help="Upload an Excel workbook (.xlsx or .xls format)"
    )
    
    if uploaded_file is not None:
        st.success(f"File uploaded: {uploaded_file.name}")
        
        # Load spell checker
        with st.spinner("Loading spell checker..."):
            spell_checker = load_spell_checker()
        
        if st.button("🔍 Check Spelling & Grammar", type="primary"):
            with st.spinner("Processing workbook... This may take a few minutes for large files."):
                try:
                    # Process the workbook
                    processed_workbook, issues_found, total_checked = process_workbook(uploaded_file, spell_checker)
                    
                    # Show summary
                    st.subheader("📊 Summary")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Cells Checked", total_checked)
                    with col2:
                        st.metric("Issues Found", len(issues_found))
                    with col3:
                        accuracy = ((total_checked - len(issues_found)) / total_checked * 100) if total_checked > 0 else 0
                        st.metric("Accuracy", f"{accuracy:.1f}%")
                    
                    # Show issues found
                    if issues_found:
                        st.subheader("⚠️ Issues Found")
                        df_issues = pd.DataFrame(issues_found)
                        st.dataframe(df_issues, use_container_width=True)
                        
                        # Create download for processed file
                        output = io.BytesIO()
                        processed_workbook.save(output)
                        output.seek(0)
                        
                        # Generate filename
                        original_name = uploaded_file.name.rsplit('.', 1)[0]
                        download_name = f"{original_name}_spell_checked.xlsx"
                        
                        st.download_button(
                            label="📥 Download Corrected File",
                            data=output.getvalue(),
                            file_name=download_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                        
                        st.info("💡 **Note:** Cells with issues are highlighted in yellow and include comments with details about the problems found.")
                        
                    else:
                        st.success("🎉 No spelling or grammar issues found!")
                        st.balloons()
                        
                        # Still offer download of original file
                        output = io.BytesIO()
                        processed_workbook.save(output)
                        output.seek(0)
                        
                        original_name = uploaded_file.name.rsplit('.', 1)[0]
                        download_name = f"{original_name}_checked.xlsx"
                        
                        st.download_button(
                            label="📥 Download File",
                            data=output.getvalue(),
                            file_name=download_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                except Exception as e:
                    st.error(f"Error processing file: {str(e)}")
                    st.error("Please make sure you've uploaded a valid Excel file.")
    
    # Instructions
    with st.expander("ℹ️ How to use this tool"):
        st.markdown("""
        1. **Upload** your Excel workbook using the file uploader above
        2. **Click** the "Check Spelling & Grammar" button to start processing
        3. **Review** the summary and list of issues found
        4. **Download** the processed file with highlighted cells
        
        **Features:**
        - ✅ Processes all sheets in the workbook
        - ✅ Highlights problematic cells in yellow
        - ✅ Adds comments to cells explaining the issues
        - ✅ Checks both spelling and grammar
        - ✅ Preserves original formatting and formulas
        - ✅ Shows detailed summary of issues found
        
        **Note:** The tool will only check cells that contain meaningful text content (not numbers, dates, or very short text).
        """)

if __name__ == "__main__":
    main()
